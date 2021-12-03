import csv
import os ,shutil
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment,Border,Side
from openpyxl.drawing.image import Image

dict1={}
dict2={}
check=[]
dict3={}
header=[]
positive=5
negative=-1
def master_roll():                                                                              #mapping roll no. with name in dict1 using master_roll.csv
    with open(os.path.join(os.getcwd()+r"/Input_files/master_roll.csv")) as f:
        reader=csv.reader(f)
        for line in reader:
            break
        for line in reader:
            dict1[line[0]]=line[1]


def responses():                                                                            #mapping roll no. with the name in dict2 with the whole list for that roll no. using responses.csv
    with open(os.path.join(os.getcwd()+r"/Input_files/responses.csv")) as f:
        reader=csv.reader(f)
        for line in reader:
            header.extend(line)
            break
        for line in reader:
            dict2[line[6]]=line

def check_ANSWER():                                                 #checking if a roll number named ANSWER exists in responses
    temp=0
    for x in dict2:
        if x=="ANSWER" :
            global check
            check.extend(dict2[x])
            temp=temp+1
    if temp==0:
        print("no roll number with ANSWER is present, Cannot Process!")
    elif temp==1:
        return
    else:
        print("more than 1 row have ANSWER as roll no.")
    return


def answer(list):                                                 # function that return a list=[no. of correct answers,no. of wrong answers,not attempted answers,no. of questions]
    questions=len(check)-7
    c=w=na=0
    stud_list=list[7:]
    correct_list=check[7:]
    for x in range(questions):
        if stud_list[x]==correct_list[x]:
            c=c+1
        elif stud_list[x]=="":
            na+=1
        else :
            w+=1
    list2=[c,w,na,questions]
    return list2




def marksheet():                                        #function that generate the marksheet of all roll no's
    if not os.path.isdir(os.getcwd()+r"/output"):
	    os.mkdir(os.getcwd()+r"/output")
    else:
        shutil.rmtree(os.getcwd()+r"/output")
        os.mkdir(os.getcwd()+r"/output")


    for x in dict1:
        wb=Workbook()
        sheet=wb.active
        sheet.title ="quiz"
        if x in dict2:                              #checking if the roll no is present in response csv if he is not means he was absent and will get a blank marksheet

            bd = Side(style='thin', color="000000")  #setting the copy of how our border will look when it will be used

            sheet.merge_cells("A5:E5")
            sheet.cell(row=5,column=1).value="Mark Sheet"
            sheet.cell(row=5,column=1).alignment = Alignment(horizontal='center')

            sheet.cell(row=6,column=1).value="Name:"
            sheet.cell(row=6,column=1).alignment = Alignment(horizontal='right')

            sheet.cell(row=6,column=2).value="{}".format(dict2[x][3])

            sheet.cell(row=6,column=4).value="Exam:"
            sheet.cell(row=6,column=4).alignment = Alignment(horizontal='right')

            sheet.cell(row=6,column=5).value="quiz"


            sheet.append(["Roll Number:",x])
            sheet.cell(row=7,column=1).alignment = Alignment(horizontal='right')

            sheet.append([])
            sheet.append(["","Right","Wrong","Not Attempt","Max"])



            question_list=answer(dict2[x])
            sheet.append(["No.",question_list[0],question_list[1],question_list[2],question_list[3]])
            sheet.append(["Marking",positive,negative,0,])
            a=question_list[0]*positive
            b=question_list[1]*negative
            sheet.append(["Total",a,b,"",str(a+b)+"/{}".format(question_list[3]*positive)])
            dict3[x]=question_list+[str(a+b)+"/{}".format(question_list[3]*positive)]
            sheet.append([""])
            sheet.append([""])

            sheet.append(["Student Ans","Correct Ans","","Student Ans","Correct Ans"])
            for j in [1,2,4,5]:
                sheet.cell(row=15,column=j).border = Border(left=bd, top=bd, right=bd, bottom=bd)

            if question_list[3]<26:          #if number of questions are less than or equal to 25
                for t in range(question_list[3]):
                    sheet.append(["{}".format(dict2[x][7+t]),"{}".format(check[7+t])])
                    sheet.cell(row=16+t,column=1).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    sheet.cell(row=16+t,column=2).border = Border(left=bd, top=bd, right=bd, bottom=bd)

            else :                          #if number of questions are more than 25
                for t1 in range(25):
                    sheet.append(["{}".format(dict2[x][7+t1]),"{}".format(check[7+t1])])
                    sheet.cell(row=16+t1,column=1).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    sheet.cell(row=16+t1,column=2).border = Border(left=bd, top=bd, right=bd, bottom=bd)

                for i in range(16,16+question_list[3]-25):
                    for j in range(4,6):
                        if j==4:
                            sheet.cell(row=i,column=j).value=dict2[x][7+25+i-16]
                        else:
                            sheet.cell(row=i,column=j).value=check[7+25+i-16]
                        sheet.cell(row=i,column=j).border = Border(left=bd, top=bd, right=bd, bottom=bd)

            for i in range(9,41):
                for j in range(1,6):
                        sheet.cell(row=i,column=j).alignment = Alignment(horizontal='center')          #aligning data of some cells to center of cell  as desired
            for i in range(9,13):
                for j in range(1,6):
                    sheet.cell(row=i,column=j).border = Border(left=bd, top=bd, right=bd, bottom=bd)    #setting borders for some cells as asked in the question


            for i in range(1,41):                   #changing the font styles like color,size,font handwriting,bold font etc of different cells as asked in the question
                for j in range(1,10):
                    if((6<=i<=7 and j==2)or(i==6 and j==5)|(i==9)or(10<=i<=12 and j==1)or(i==15)):
                        sheet.cell(row=i,column=j).font = Font(bold=True,size="12",name="Century")
                    elif((10<=i<=12 and j==2)or((16<=i<=40 and j in[1,4])and(sheet.cell(row=i,column=j).value==sheet.cell(row=i,column=j+1).value))):#for green color font
                        sheet.cell(row=i,column=j).font = Font(size="12",name="Century",color = "00008000")
                    elif((10<=i<=12 and j==3)or((16<=i<=40 and j in[1,4])and(sheet.cell(row=i,column=j).value!=sheet.cell(row=i,column=j+1).value))):#for red color font
                        sheet.cell(row=i,column=j).font = Font(size="12",name="Century",color = "00FF0000")
                    elif((i==12 and j==5)or(16<=i<=40 and j in [2,5])):
                        sheet.cell(row=i,column=j).font = Font(size="12",name="Century",color = "000000FF")                                     #for blue color font
                    elif(i==5 and j==1):
                        sheet.cell(row=i,column=j).font = Font(bold=True,underline='single',size="18",name="Century")
                    else:
                        sheet.cell(row=i,column=j).font = Font(size="12",name="Century")

            for column in range(1,6):
                sheet.column_dimensions[chr(column+64)].width = 24.4                                    #increasing width of first 5 colums to adjust the pic of iitp logo

            for i in range(1,5):                                    #increasing height of first 5 rows to adjust the pic of iitp logo
                sheet.row_dimensions[i].height = 22.7

            my_png = Image(os.getcwd()+r"/iitp_logo.png")                   #inserting the pic of iit patna logo(filename-iitp_logo.png) to the marksheet
            sheet.add_image(my_png, 'A1')

        wb.save(os.getcwd()+r"/xyz/"+"{}.xlsx".format(x))               #saving the marksheet of a roll no.



def concise_marksheet():
    with open(os.getcwd()+r"/concise_marksheet.csv",'w',newline='') as file:
        writer=csv.writer(file)
        line=[]
        line.extend(header)
        for x in range(len(line)):
            if x>6 :
                    line[x]="Unnamed: {}".format(x)
            elif x==2:
                line[x]="Google_Score"
        writer.writerow(line[0:6]+["Score_After_Negative"]+line[6:]+["statusAns"])
        for x in dict2:
            print(dict2[x][0:6]+[dict3[x][4]]+dict2[x][6:]+dict3[x][0:3])
            writer.writerow(dict2[x][0:6]+[dict3[x][4]]+dict2[x][6:]+[dict3[x][0:3]])






master_roll()
responses()
check_ANSWER()
marksheet()
concise_marksheet()
print("Running complete")
#print(dict3)
