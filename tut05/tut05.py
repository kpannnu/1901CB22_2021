def generate_marksheet():
    import os
    import csv
    from openpyxl import Workbook
    from openpyxl import load_workbook
    try:
        os.mkdir(r"./output")
    except:
        path=r"./output"

    dict={"AA":10,"AA*":10,"AB":9,"AB*":9,"BB":8,"BB*":8,"BC":7,"BC*":7,"CC":6,"CC*":6,"CD":5,"CD*":5,"DD":4,"DD*":4,"F":0,"F*":0,"I":0,"I*":0}#mapping the grades to grade equivalent
    dict1,dict2,dict3={},{},{}
    i=2
    with open("grades.csv","r") as f1:   #opening grades.csv file and itearting it then 
        grade_file_reader=csv.reader(f1)
        next(grade_file_reader)
        for x in grade_file_reader:
            dict1[i]=x                    #converting grades.csv file into dictionary
            i=i+1
                  
    with open("subjects_master.csv","r") as f2: #opening subjects_master.csv file and itearting it then 
        subject_file_reader=csv.reader(f2)
        for y in subject_file_reader:
            dict2[y[0]]=[y[1],y[2]]        #converting subjects_master.csv file into dictionary

    with open("names-roll.csv") as f3:       #opening names-roll.csv file and itearting it then    
        name_file_reader=csv.reader(f3)
        for z in name_file_reader:
            dict3[z[0]]=z[1]               #converting names-roll.csv file into dictionary

    def fun(roll):# function for calculation of all information to form Overall sheet and the each semester sheets using student roll no 
        wb=Workbook()
        count=credit=temp=0
        k=1                #variable made to check when semester is changing for same roll no. in the iteration
        list=[]
        list2=[]
        for x in dict1:    # iterating dictionary of grades.csv file
            if(dict1[x][0]==roll):
                if(int(dict1[x][1])!=k):
                    list.append([k,credit,temp/credit]) # credit=semester wise credit taken, temp=credit*grade equivalent, temp/credit = SPI of a particular sem
                    credit=temp=0
                if "Sem{}".format(dict1[x][1]) not in wb.sheetnames: # information for a particular Sem of the roll no. provided as argument in the function 
                    wb.create_sheet(index=int(dict1[x][1]),title="Sem{}".format(dict1[x][1]))
                    wb["Sem{}".format(dict1[x][1])].append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                sheet=wb["Sem{}".format(dict1[x][1])]
                sheet.append([sheet.max_row,dict1[x][2],dict2[dict1[x][2]][0],dict2[dict1[x][2]][1],int(dict1[x][3]),dict1[x][5],dict1[x][4]])
                credit=credit+int(dict1[x][3])
                temp=temp+int(dict1[x][3])*dict[dict1[x][4].strip()] 
                k=int(dict1[x][1])
                count=count+1
            
            if(dict1[x][0]!=roll or x==len(dict1)):
                if(count!=0):
                    list.append([k,credit,temp/credit])
                    wb.create_sheet(index=0,title="Overall")
                    sheet=wb["Overall"]
                    sheet.append(["Roll No.",roll])
                    sheet.append(["Name of Student ",dict3[roll]])
                    sheet.append(["Discipline",roll[4:6]])

                    sheet.append(["Semester No."]+[x[0] for x in list])
                    sheet.append(["Semester wise Credit Taken"]+[x[1] for x in list])
                    sheet.append(["SPI"]+[round(x[2],2) for x in list])

                    def foo(x,list=list):           # function to calculate total credits taken till the xth semester given as argument
                        total_credit=0
                        for y in list[:x]:
                            total_credit=total_credit+y[1]
                        return total_credit
                    sheet.append(["Total Credits Taken"]+[foo(x[0]) for x in list])

                    def cpi(x,list=list):           # function to calculate CPI till xth semester given as argument with roundoff of 2 decimal digits
                        total_credit=0
                        total_temp=0
                        for y in list[:x]:
                            total_credit=total_credit+y[1]
                            total_temp=total_temp+y[1]*y[2]
                        cpi=round(total_temp/total_credit,2)
                        return cpi

                    sheet.append(["CPI"]+[cpi(x[0]) for x in list])
                    wb.remove(wb["Sheet"])
                    wb.save(".\output\{}.xlsx".format(roll))  #saving the .xlsx file for a Student 
                    return
    for z in dict3:  # iterating the dictionary of names-roll.csv file to get unique roll no. every time
        fun(z)       # calling the function to calculate the data of a Student by passing his unique roll no. as argument
                        
    return

generate_marksheet() #calling function to generate marksheet






