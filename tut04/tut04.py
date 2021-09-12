import csv             #importing modules
from openpyxl import Workbook
from openpyxl import load_workbook

def output_by_subject():      #defining function by subject
    with open("regtable_old.csv","r") as f:
        reader=csv.reader(f)
        for x in reader:
            header=[x[0],x[1],x[3],x[8]]
            break
        
        for x in reader:
            try:
                wb=load_workbook(r"./output_by_subject/{}.xlsx".format(x[3]))
                sheet=wb.active
                list=[x[0],x[1],x[3],x[8]]
                sheet.append(list)
                wb.save(r"./output_by_subject/{}.xlsx".format(x[3]))
            except:
                wb=Workbook()
                sheet=wb.active
                list=[x[0],x[1],x[3],x[8]]
                sheet.append(header)
                sheet.append(list)
                wb.save('./output_by_subject/{}.xlsx'.format(x[3]))
            
    return

def output_individual_roll():  #defining function for individual roll 
    with open("regtable_old.csv","r") as f:
        reader=csv.reader(f)
        for x in reader:
            header=[x[0],x[1],x[3],x[8]]
            break
        
        for x in reader:
            try:
                wb=load_workbook(r"./output_individual_roll/{}.xlsx".format(x[0]))
                sheet=wb.active
                list=[x[0],x[1],x[3],x[8]]
                sheet.append(list)
                wb.save(r"./output_individual_roll/{}.xlsx".format(x[0]))
            except:
                wb=Workbook()
                sheet=wb.active
                list=[x[0],x[1],x[3],x[8]]
                sheet.append(header)
                sheet.append(list)
                wb.save('./output_individual_roll/{}.xlsx'.format(x[0]))
            
    return

output_by_subject()  #calling function
output_individual_roll() #calling function

#Note:-The program took around 3-4 minutes for execution in my system




